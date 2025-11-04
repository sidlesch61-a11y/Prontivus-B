"""
Reporting services: generate PDF and Excel exports for analytics with simple clinic branding.
"""

from io import BytesIO
from typing import Any, Dict

from reportlab.lib.pagesizes import A4
from reportlab.pdfgen import canvas
from reportlab.lib.units import cm
from openpyxl import Workbook
import os


def generate_analytics_pdf(title: str, data: Dict[str, Any], clinic_name: str = "Prontivus") -> bytes:
    buffer = BytesIO()
    c = canvas.Canvas(buffer, pagesize=A4)
    width, height = A4

    # Branding header
    try:
        logo_path = os.getenv("PRONTIVUS_LOGO_PATH", "public/Logo/Prontivus Horizontal Transparents.png")
        if os.path.exists(logo_path):
            # drawImage(x, y, width=None, height=None, preserveAspectRatio=True, mask=None)
            # Height 60px ~= 2.12 cm at 72 dpi; we set height in points directly
            c.drawImage(logo_path, 2 * cm, height - 2.5 * cm, height=60, preserveAspectRatio=True, mask='auto')
            text_x = 2 * cm + 6 * cm
        else:
            text_x = 2 * cm
    except Exception:
        text_x = 2 * cm
    c.setFont("Helvetica-Bold", 16)
    c.drawString(text_x, height - 2 * cm, clinic_name)
    c.setFont("Helvetica", 12)
    c.drawString(text_x, height - 2.8 * cm, title)
    # Slogan
    c.setFont("Helvetica-Oblique", 10)
    c.drawString(text_x, height - 3.4 * cm, "Prontivus — Cuidado inteligente")

    y = height - 4.6 * cm
    c.setFont("Helvetica", 10)
    for key, value in data.items():
        text = f"{key}: {value}"
        c.drawString(2 * cm, y, text[:110])
        y -= 0.6 * cm
        if y < 2 * cm:
            c.showPage()
            y = height - 2 * cm

    c.showPage()
    c.save()
    buffer.seek(0)
    return buffer.read()


def generate_analytics_excel(title: str, data: Dict[str, Any]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = (title or "Report")[:31]

    row = 1
    for key, value in data.items():
        ws.cell(row=row, column=1, value=str(key))
        ws.cell(row=row, column=2, value=str(value))
        row += 1

    out = BytesIO()
    wb.save(out)
    out.seek(0)
    return out.read()


